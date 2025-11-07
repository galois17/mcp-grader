from fastmcp import FastMCP
import pandas as pd
import json
import os
import sys
import uuid
import logging
import glob
import time
from datetime import datetime
from docx import Document
import boto3
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from prompt_templates import build_excel_prompt, build_word_prompt
from typing import Dict, Any, Optional, List
from decimal import Decimal
from helpers import *
import math
import re

logging.basicConfig(stream=sys.stderr, level=logging.INFO)
logger = logging.getLogger(__name__)

# Global Config
app_id = os.environ.get("__app_id", "default-app-id")
user_id = os.environ.get("__user_id", "default-user-id")
BEDROCK_REGION = "us-east-2"
MODEL_ID = "arn:aws:bedrock:us-east-2:073392940626:inference-profile/us.amazon.nova-micro-v1:0"

# Initialize AWS Clients
try:
    boto3_session = boto3.Session(region_name=BEDROCK_REGION)
    bedrock = boto3_session.client("bedrock-runtime")
    dynamodb = boto3_session.resource("dynamodb")
    print("AWS Bedrock & DynamoDB initialized.")
except Exception as e:
    print(f"ERROR initializing AWS clients: {e}", file=sys.stderr)
    bedrock = None
    dynamodb = None

# DynamoDB Table Setup
TEMPLATE_TABLE_NAME = "AssignmentTemplatesTable"
GRADED_TABLE_NAME = "GradedAssignmentsTable"

def ensure_tables_exist():
    existing_tables = [t.name for t in dynamodb.tables.all()]
    if TEMPLATE_TABLE_NAME not in existing_tables:
        dynamodb.create_table(
            TableName=TEMPLATE_TABLE_NAME,
            KeySchema=[{"AttributeName": "PK", "KeyType": "HASH"}],
            AttributeDefinitions=[{"AttributeName": "PK", "AttributeType": "S"}],
            BillingMode="PAY_PER_REQUEST",
        )
        dynamodb.Table(TEMPLATE_TABLE_NAME).wait_until_exists()
        print(f"Created {TEMPLATE_TABLE_NAME} (on-demand)")
    if GRADED_TABLE_NAME not in existing_tables:
        dynamodb.create_table(
            TableName=GRADED_TABLE_NAME,
            KeySchema=[{"AttributeName": "PK", "KeyType": "HASH"}],
            AttributeDefinitions=[{"AttributeName": "PK", "AttributeType": "S"}],
            BillingMode="PAY_PER_REQUEST",
        )
        dynamodb.Table(GRADED_TABLE_NAME).wait_until_exists()
        print(f"Created {GRADED_TABLE_NAME} (on-demand)")

if dynamodb:
    ensure_tables_exist()
    template_table = dynamodb.Table(TEMPLATE_TABLE_NAME)
    graded_table = dynamodb.Table(GRADED_TABLE_NAME)
else:
    template_table = graded_table = None

# Initialize MCP
mcp = FastMCP("SimpleGraderServer")

def _invoke_bedrock_extraction(prompt: str) -> Dict[str, Any]:
    if not bedrock:
        raise Exception("Bedrock client not initialized.")
    body = json.dumps({
        "messages": [{"role": "user", "content": [{"text": prompt}]}],
        "inferenceConfig": {"maxTokens": 900, "temperature": 0.0, "topP": 1.0},
    })
    response = bedrock.invoke_model(
        modelId=MODEL_ID, contentType="application/json", accept="application/json", body=body
    )
    response_body = json.loads(response["body"].read())
    output_text = (
        response_body.get("output", {}).get("message", {}).get("content", [{}])[0].get("text", "")
    ).strip()
    if not output_text:
        raise Exception("LLM returned empty content.")
    if "```" in output_text:
        if "```json" in output_text:
            output_text = output_text.split("```json", 1)[1].split("```", 1)[0].strip()
        else:
            output_text = output_text.split("```", 1)[1].split("```", 1)[0].strip()
    if not (output_text.startswith("{") and output_text.rstrip().endswith("}")):
        js, je = output_text.find("{"), output_text.rfind("}")
        if js != -1 and je != -1:
            output_text = output_text[js : je + 1]
    extracted_json =  json.loads(output_text)
    # Ensure each item includes reason, even if missing
    for item in extracted_json.get("items", []):
        if "reason" not in item:
            item["reason"] = "N/A"
    return extracted_json

# List Tools
@mcp.tool()
def list_all_tools() -> List[Dict[str, Any]]:
    tools = []
    for tool in mcp.registry.tools.values():
        params = []
        if hasattr(tool, "signature"):
            for p in tool.signature.parameters.values():
                params.append(
                    {"name": p.name, "annotation": str(p.annotation), "default": getattr(p, "default", None)}
                )
        tools.append({"name": tool.name, "description": getattr(tool, "description", ""), "parameters": params})
    return tools

# Tool: Save Assignment Templat
@mcp.tool()
def set_assignment_structure(    
    file_path: str, 
    question_col: Optional[str] = None, 
    answer_col: Optional[str] = None,
    id_col: Optional[str] = None) -> Dict[str, Any]:
    """
    Reads an answer key Excel file, generates an extraction prompt,
    and saves this template to DynamoDB, returning a unique ID.
    This template is used to process student submissions.
    """
    if not template_table:
        return {"status": "error", "message": "DynamoDB not initialized."}
    try:
        table_text = read_to_text(file_path)

        if file_path.lower().endswith(".xlsx"):
            prompt = build_excel_prompt(table_text)
        elif file_path.lower().endswith(".docx"):
            prompt = build_word_prompt(table_text)
        else:
            raise ValueError(f"Unsupported file type for template: {file_path}")

        template_id = str(uuid.uuid4())
        item = {
            "PK": template_id,
            "app_id": app_id,
            "user_id": user_id,
            "file_path": file_path,
            "filename": os.path.basename(file_path),
            "extraction_prompt": prompt,
            "structure_details": None,
            "created_at": datetime.now().isoformat(),
            "status": "Template Saved",
        }
        template_table.put_item(Item=item)
        return {"status": "success", "template_id": template_id, "message": "Template saved."}
    except Exception as e:
        logger.exception("Error in set_assignment_structure")
        return {"status": "error", "message": str(e)}

# Tool: Analyze Template
@mcp.tool()
def analyze_assignment_structure(template_id: str) -> Dict[str, Any]:
    """
    Retrieves the prompt for an answer key template from Firestore,
    invokes the Bedrock LLM to extract the structure,
    and updates the Firestore document with the extracted JSON details.
    """
    if not bedrock or not template_table:
        return {"status": "error", "message": "Clients not initialized."}
    try:
        res = template_table.get_item(Key={"PK": template_id})
        item = res.get("Item")
        if not item:
            return {"status": "error", "message": "Template not found."}
        extracted = _invoke_bedrock_extraction(item["extraction_prompt"])
        item["structure_details"] = {**extracted, "filename": item.get("filename")}
        item["status"] = "Structure Analyzed"
        item["updated_at"] = datetime.now().isoformat()
        template_table.put_item(Item=item)
        
        return {
            "status": "success",
            "message": "Template analyzed.",
            "template_id": template_id,
            "filename": item.get("filename"),
            "llm_data_summary": {
                "total_points_cell": extracted.get("total_points_cell"),
                "items_count": len(extracted.get("items", [])),
            },
            "llm_output": {**extracted, "filename": item.get("filename")}
        }
    except Exception as e:
        logger.exception("Error in analyze_assignment_structure")
        return {"status": "error", "message": str(e)}

# Tool: Upload Student Submission
@mcp.tool()
def upload_item_to_grade_key_file(template_id: str, file_path: str) -> Dict[str, Any]:
    """
    Uploads a student's assignment file and prepares it for grading against an existing
    analyzed answer key template.

    This function performs the following steps:
      1. Retrieves the analyzed answer key (template) from the DynamoDB table
         using the provided `template_id`.
      2. Reads the student's submission file (Excel or Word) and constructs a
         specific extraction prompt by injecting the student's data into the
         base extraction prompt used for the answer key.
      3. Saves a new record into the GradedAssignmentsTable, linking this
         student submission to its corresponding template.
      4. Returns a unique `student_assignment_id` that can be used by
         downstream tools such as `analyze_item_to_grade_structure()` and
         `grade_all_submissions()`.

    The stored record includes:
      - Template ID and user ID linkage
      - File path and extracted prompt
      - Status tracking ("Pending Extraction")
      - Placeholder for grading and extracted data fields
    """
    return _upload_item_to_grade_key_file(template_id, file_path)

def _upload_item_to_grade_key_file(template_id: str, file_path: str) -> Dict[str, Any]:
    if not graded_table or not template_table:
        return {"status": "error", "message": "DynamoDB not initialized."}
    try:
        res = template_table.get_item(Key={"PK": template_id})
        template_data = res.get("Item")
        if not template_data:
            return {"status": "error", "message": f"Template {template_id} not found."}
        extraction_prompt = template_data.get("extraction_prompt")
        if not extraction_prompt:
            return {"status": "error", "message": "Template missing prompt."}
        
        student_table_text = read_to_text(file_path)
        student_prompt = extraction_prompt.split("### Spreadsheet Data:")[0] + \
                         f"### Spreadsheet Data:\n{student_table_text}"
        student_id = str(uuid.uuid4())
        filename = os.path.basename(file_path)
        item = {
            "PK": student_id,
            "template_id": template_id,
            "user_id": user_id,
            "student_file_path": file_path,
            "filename": filename,
            "extraction_prompt_used": student_prompt,
            "extracted_data": None,
            "status": "Pending Extraction",
            "grade": None,
            "created_at": datetime.now().isoformat(),
        }
        graded_table.put_item(Item=item)
        return {"status": "success", "student_assignment_id": student_id, "filename": filename}
    except Exception as e:
        logger.exception("Error in upload_item_to_grade_key_file")
        return {"status": "error", "message": str(e)}

# Tool: Analyze Student Submission
@mcp.tool()
def analyze_item_to_grade_structure(student_assignment_id: str) -> Dict[str, Any]:
    return  _analyze_item_to_grade_structure(student_assignment_id)


def _analyze_item_to_grade_structure(student_assignment_id: str) -> Dict[str, Any]:
    if not bedrock or not graded_table:
        return {"status": "error", "message": "Clients not initialized."}
    try:
        res = graded_table.get_item(Key={"PK": student_assignment_id})
        item = res.get("Item")
        if not item:
            return {"status": "error", "message": "Student submission not found."}
        extracted = _invoke_bedrock_extraction(item["extraction_prompt_used"])
        item["extracted_data"] = {**extracted, "filename": item.get("filename")}
        item["status"] = "Extracted"
        item["updated_at"] = datetime.now().isoformat()
        graded_table.put_item(Item=item)
        return {
            "status": "success",
            "student_assignment_id": student_assignment_id,
            "filename": item.get("filename"),
            "llm_data_summary": {
                "total_points_cell": extracted.get("total_points_cell"),
                "items_count": len(extracted.get("items", [])),
            },
            "llm_output": {**extracted, "filename": item.get("filename")}
        }
    except Exception as e:
        logger.exception("Error in analyze_item_to_grade_structure")
        return {"status": "error", "message": str(e)}

# Tool: Batch Process Student Folder
@mcp.tool()
def batch_process_student_folder(template_id: str, folder_path: str) -> Dict[str, Any]:
    if not os.path.isdir(folder_path):
        return {"status": "error", "message": f"Folder not found: {folder_path}"}
    if not template_table or not graded_table:
        return {"status": "error", "message": "DynamoDB not initialized."}

    results = []
    excel_files = sorted(
        glob.glob(os.path.join(folder_path, "*.xlsx"))
        + glob.glob(os.path.join(folder_path, "*.xls")) 
        + glob.glob(os.path.join(folder_path, "*.doc"))
        + glob.glob(os.path.join(folder_path, "*.docx"))
    )

    if not excel_files:
        return {"status": "error", "message": "No Excel files found in folder."}

    for file_path in excel_files:
        try:
            # upload_res = upload_item_to_grade_key_file(template_id=template_id, file_path=file_path)
            upload_res = _upload_item_to_grade_key_file(template_id=template_id, file_path=file_path)
       
            if upload_res.get("status") != "success":
                results.append({
                    "file": os.path.basename(file_path),
                    "status": "upload_failed",
                    "message": upload_res.get("message", "Unknown error")
                })
                continue
            student_id = upload_res["student_assignment_id"]
            # analyze_res = analyze_item_to_grade_structure(student_id)
            analyze_res = _analyze_item_to_grade_structure(student_id)
            results.append({
                "file": os.path.basename(file_path),
                "student_assignment_id": student_id,
                "filename": os.path.basename(file_path),
                "status": analyze_res.get("status"),
                "message": analyze_res.get("message", ""),
                "items_count": analyze_res.get("llm_data_summary", {}).get("items_count", 0)
            })
            time.sleep(1.5)
        except Exception as e:
            logger.exception(f"Error processing {file_path}")
            results.append({"file": os.path.basename(file_path), "status": "error", "message": str(e)})

    success_count = sum(1 for r in results if r["status"] == "success")

    return {
        "status": "completed",
        "summary": {"total_files": len(excel_files), "successfully_processed": success_count},
        "results": results,
        "student_assignment_ids": [r["student_assignment_id"] for r in results if "student_assignment_id" in r]
    }

def _convert_floats(obj: Any) -> Any:
    """
    Recursively walk a Python object (dict/list/scalar)
    and convert all floats to Decimal for DynamoDB.
    """
    if isinstance(obj, float):
        # preserve string precision
        return Decimal(str(obj))
    elif isinstance(obj, list):
        return [_convert_floats(v) for v in obj]
    elif isinstance(obj, dict):
        return {k: _convert_floats(v) for k, v in obj.items()}
    else:
        return obj
 

@mcp.tool()
def grade_all_submissions(template_id: str) -> Dict[str, Any]:
    """
    Compare each student's extracted data against the analyzed template.
    Produces a styled Excel workbook with:
      1. 'Detailed Breakdown' – full per-question grading with colors
      2. 'Condensed Summary' – per-student overview with emojis, lists, and totals.
    """
    # Load template
    t_doc = template_table.get_item(Key={"PK": template_id}).get("Item")
    if not t_doc:
        return {"status": "error", "message": f"Template {template_id} not found."}

    structure = t_doc.get("structure_details")
    if not structure or not isinstance(structure, dict):
        return {"status": "error", "message": "Template has no analyzed structure_details."}

    key_items = structure.get("items")
    if not key_items:
        return {"status": "error", "message": "Template structure missing 'items' list."}

    # Pull student submissions
    scan = graded_table.scan()
    students = [s for s in scan.get("Items", []) if s.get("status") == "Extracted"]
    if not students:
        return {"status": "error", "message": "No extracted student submissions found."}


    wb = Workbook()
    ws = wb.active
    ws.title = "Detailed Breakdown"
    ws_summary = wb.create_sheet(title="Condensed Summary")

    # Shared styles
    header_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    header_font = Font(bold=True, color="000000")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # Sheet 1 headers
    headers = [
        "filename", "question", "correct_answer", "student_answer",
        "confidence", "reason", "points_awarded", "total_points"
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Sheet 2 headers
    summary_headers = [
        "filename",
        "low_confidence_flag",
        "low_confidence_answers",
        "wrong_answers",
        "reason_summary",
        "total_points"
    ]
    ws_summary.append(summary_headers)
    for cell in ws_summary[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Colors
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    red_fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
    gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    # Grading + data collection
    summary_rows = []

    for s in students:
        s_items = s.get("extracted_data", {}).get("items", [])
        if not s_items:
            continue

        filename = s.get("filename", s.get("student_file_path", "unknown"))
        total_points = 0.0
        low_conf_flag = False
        low_conf_answers = []
        wrong_answers = []
        reasons = set()

        for k, q_key in enumerate(key_items):
            correct = str(q_key.get("answer", "")).strip().lower()
            pts_raw = q_key.get("points", "0").replace("pt", "").replace("s", "")
            try:
                pts = float(pts_raw)
            except ValueError:
                pts = 0.0

            if k < len(s_items):
                q_student = s_items[k]
                student_ans = str(q_student.get("answer", "")).strip().lower()
                conf = q_student.get("confidence", "unknown")
                reason = q_student.get("reason", "N/A")
            else:
                student_ans, conf, reason = "", "missing", "N/A"

            # Grading logic 
            awarded = 0.0
            is_wrong = False
            if is_number(student_ans) and is_number(correct):
                if numbers_close(student_ans, correct, tol_decimals=2):
                    awarded = pts
                else:
                    diff = abs(float(student_ans) - float(correct))
                    rel_err = diff / (abs(float(correct)) + 1e-9)
                    if rel_err < 0.05:
                        awarded = pts / 2.0
                    else:
                        is_wrong = True
            elif student_ans.strip() == correct.strip():
                awarded = pts
            elif conf == "low" and correct and correct[:3] in student_ans:
                awarded = pts / 2.0
            else:
                is_wrong = True

            total_points += awarded

            # Track summary info
            if conf.lower() == "low":
                low_conf_flag = True
                low_conf_answers.append(student_ans or "(blank)")
            if is_wrong:
                wrong_answers.append(student_ans or "(blank)")
            if reason and reason != "N/A":
                reasons.add(reason)

            # Append detailed row
            ws.append([
                filename, q_key.get("question", ""), correct, student_ans,
                conf, reason, awarded, ""
            ])
            # Color coding
            fill = green_fill if awarded == pts else yellow_fill if awarded > 0 else red_fill
            for cell in ws[ws.max_row]:
                cell.fill = fill
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Total row in sheet 1
        ws.append([filename, "TOTAL", "", "", "", "", "", total_points])
        for cell in ws[ws.max_row]:
            cell.fill = gray_fill
            cell.font = Font(bold=True)
            cell.border = thin_border

        # Condensed summary entry 
        summary_rows.append({
            "filename": filename,
            "low_conf_flag": low_conf_flag,
            "low_conf_answers": ", ".join(low_conf_answers) if low_conf_answers else "N/A",
            "wrong_answers": ", ".join(wrong_answers) if wrong_answers else "N/A",
            "reason_summary": "; ".join(sorted(reasons)) if reasons else "N/A",
            "total_points": round(total_points, 2)
        })

        # Update DynamoDB
        s["grade"] = {"total": total_points}
        s["status"] = "Graded"
        safe_item = _convert_floats(s)
        graded_table.put_item(Item=safe_item)

    #  Write condensed summary 
    for entry in summary_rows:
        emoji = "😟" if entry["low_conf_flag"] else "😊"
        ws_summary.append([
            entry["filename"],
            emoji,
            entry["low_conf_answers"],
            entry["wrong_answers"],
            entry["reason_summary"],
            entry["total_points"]
        ])
        fill = red_fill if entry["low_conf_flag"] else green_fill
        for cell in ws_summary[ws_summary.max_row]:
            # Only color the emoji column (B)
            if cell.column_letter == "B":
                cell.fill = fill
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Autosize all columns
    for sheet in [ws, ws_summary]:
        for col in sheet.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
            sheet.column_dimensions[col_letter].width = min(max_len + 2, 50)

    
    output_filename = f"grading_results_{template_id}.xlsx"
    output_path = os.path.join(os.getcwd(), output_filename)
    wb.save(output_path)

    return {
        "status": "success",
        "message": f"Grading complete. XLSX written to {output_path}",
        "xlsx_path": output_path
    }


# Tool: Retrieve All Student Assignments
@mcp.tool()
def get_all_student_assignments() -> Dict[str, Any]:
    return _get_all_student_assignments_core()

def _get_all_student_assignments_core() -> Dict[str, Any]:
    if not graded_table:
        return {"status": "error", "message": "DynamoDB not initialized."}
    try:
        all_items = []
        scan_kwargs = {}
        while True:
            response = graded_table.scan(**scan_kwargs)
            all_items.extend(response.get("Items", []))
            if "LastEvaluatedKey" not in response:
                break
            scan_kwargs["ExclusiveStartKey"] = response["LastEvaluatedKey"]

        assignments = []
        for item in all_items:
            assignments.append({
                "student_assignment_id": item["PK"],
                "template_id": item.get("template_id"),
                "file_path": item.get("student_file_path"),
                "filename": item.get("filename"),
                "status": item.get("status"),
                "grade": item.get("grade"),
                "created_at": item.get("created_at"),
                "items_extracted": len(item.get("extracted_data", {}).get("items", []))
                    if isinstance(item.get("extracted_data"), dict)
                    else None
            })

        return {
            "status": "success",
            "summary": {
                "total_assignments": len(assignments),
                "graded": sum(1 for a in assignments if a.get("status") == "Graded"),
                "extracted": sum(1 for a in assignments if a.get("status") == "Extracted"),
                "pending": sum(1 for a in assignments if a.get("status") == "Pending Extraction"),
            },
            "assignments": assignments
        }
    except Exception as e:
        logger.exception("Error in get_all_student_assignments")
        return {"status": "error", "message": str(e)}

# Run MCP Server
if __name__ == "__main__":
    print("!! Starting MCP Grading Server (DynamoDB on-demand)...")
    mcp.run(transport="http", host="0.0.0.0", port=8000)
